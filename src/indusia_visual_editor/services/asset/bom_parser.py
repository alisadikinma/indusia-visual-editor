"""BOM (Bill of Materials) parser for Excel + CSV inputs.

Public surface:
- `parse_bom(file_bytes, filename) -> list[BomItemDraft]`
- `BomItemDraft` — pre-DB pydantic representation
- `BomParseError` — typed error with Bahasa Indonesia message

Behavior:
- Header is auto-detected within the first 5 rows by counting recognized
  column names. The first row with ≥2 known headers wins.
- Fuzzy column matching: `designator|reference|ref|comp` → designator,
  `value|val` → value, `package|footprint|fp` → package, `qty|quantity|qnty`
  → qty. All matching is case-insensitive substring.
- Multi-designator rows like `R1, R2, R3` or `C4;C5` expand to N items
  sharing value/package/extra. The plan calls this out explicitly (§5.3).
- Unrecognized columns are preserved in `extra: dict[str, str]` (no silent
  data loss).
- Empty file or missing designator column raises `BomParseError` with a
  Bahasa-Indonesia-friendly message (the UI surface for factory users).
"""

from __future__ import annotations

import csv
import io
import re
from pathlib import Path

from openpyxl import load_workbook
from pydantic import BaseModel

from indusia_visual_editor.services.asset.mi_classifier import classify
from indusia_visual_editor.utils.logging_config import get_logger

logger = get_logger(__name__)


# Fuzzy column synonyms — substring match, case-insensitive.
COLUMN_SYNONYMS: dict[str, tuple[str, ...]] = {
    "designator": ("designator", "reference", "ref", "comp"),
    "value": ("value", "val"),
    "package": ("package", "footprint", "fp"),
    "qty": ("qty", "quantity", "qnty"),
}

# Header detection scans this many leading rows.
HEADER_SCAN_WINDOW = 5

# Multi-designator expansion: split on `,` or `;` (with optional whitespace).
DESIGNATOR_SPLIT = re.compile(r"[,;]\s*")


class BomParseError(ValueError):
    """Raised when the BOM cannot be parsed. Message MUST be user-friendly
    Bahasa Indonesia so it can be surfaced directly in the UI."""


class BomItemDraft(BaseModel):
    """Pre-DB representation. The route layer converts these into
    `db.models.BomItem` rows after persistence."""

    designator: str
    value: str | None = None
    package: str | None = None
    qty: int | None = None
    extra: dict[str, str] | None = None
    # Heuristic hints set by mi_classifier — UI badge + smart-select only,
    # never auto-decides inspect_scope (Phase 2.2b).
    mi_likely: bool = False
    component_type: str | None = None


def _classify_header(cell: str) -> str | None:
    if not cell:
        return None
    lower = cell.strip().lower()
    for canonical, synonyms in COLUMN_SYNONYMS.items():
        for syn in synonyms:
            if syn in lower:
                return canonical
    return None


def _detect_header(rows: list[list[str]]) -> tuple[int, dict[int, str]]:
    """Return (header_row_index, column_index → canonical_name).
    Raises BomParseError when no row in the scan window has the designator
    column."""
    best_idx = -1
    best_mapping: dict[int, str] = {}
    for idx, row in enumerate(rows[:HEADER_SCAN_WINDOW]):
        mapping: dict[int, str] = {}
        for col_idx, cell in enumerate(row):
            canonical = _classify_header(str(cell) if cell is not None else "")
            if canonical is not None:
                mapping[col_idx] = canonical
        if "designator" in mapping.values() and len(mapping) > len(best_mapping):
            best_idx = idx
            best_mapping = mapping
    if best_idx < 0 or "designator" not in best_mapping.values():
        scanned = [str(c) for c in (rows[0] if rows else []) if c]
        raise BomParseError(
            "Kolom 'designator' tidak ditemukan dalam BOM. "
            f"Kolom yang terbaca: {scanned!r}. "
            "Pastikan ada kolom 'Designator', 'Reference', atau 'Ref' di header."
        )
    return best_idx, best_mapping


def _parse_qty(raw: object) -> int | None:
    if raw is None or raw == "":
        return None
    try:
        return int(float(str(raw).strip()))
    except (ValueError, TypeError):
        return None


def _row_to_drafts(
    row: list[object], mapping: dict[int, str], header_row: list[str]
) -> list[BomItemDraft]:
    """Build BomItemDrafts from one body row. Multi-designator cells
    (`R1, R2, R3`) expand to N drafts sharing value/package/extra."""
    fields: dict[str, object] = {}
    extra: dict[str, str] = {}
    for col_idx, cell in enumerate(row):
        canonical = mapping.get(col_idx)
        header_name = (
            str(header_row[col_idx]) if col_idx < len(header_row) else f"col_{col_idx}"
        )
        if canonical is None:
            if cell is not None and str(cell).strip() != "" and header_name:
                extra[header_name] = str(cell)
            continue
        if canonical == "qty":
            fields["qty"] = _parse_qty(cell)
        else:
            value = None if cell is None else str(cell).strip()
            fields[canonical] = value or None

    raw_designator = fields.get("designator")
    if not raw_designator:
        return []

    designators = [d.strip() for d in DESIGNATOR_SPLIT.split(str(raw_designator)) if d.strip()]
    if not designators:
        return []

    drafts: list[BomItemDraft] = []
    for d in designators:
        pkg = fields.get("package") or ""
        val = fields.get("value") or ""
        cls = classify(package=str(pkg), value=str(val), designator=d)
        drafts.append(
            BomItemDraft(
                designator=d,
                value=fields.get("value"),  # type: ignore[arg-type]
                package=fields.get("package"),  # type: ignore[arg-type]
                qty=fields.get("qty"),  # type: ignore[arg-type]
                extra=extra or None,
                mi_likely=cls.mi_likely,
                component_type=cls.component_type,
            )
        )
    return drafts


def _load_xlsx(file_bytes: bytes) -> list[list[object]]:
    if not file_bytes:
        raise BomParseError("File BOM kosong.")
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises various Invalid* exceptions
        raise BomParseError(f"File BOM .xlsx tidak bisa dibaca: {exc}") from exc
    ws = wb.active
    rows: list[list[object]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    wb.close()
    return rows


def _load_csv(file_bytes: bytes) -> list[list[object]]:
    if not file_bytes:
        raise BomParseError("File BOM kosong.")
    text = file_bytes.decode("utf-8-sig", errors="replace")
    # Sniffer for delimiter — most BOMs are comma; some Excel exports use ;
    sample = text[:1024]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect=dialect)
    try:
        return [list(row) for row in reader]
    except csv.Error as exc:
        raise BomParseError(
            "File BOM .csv tidak bisa dibaca — kemungkinan ada newline di dalam "
            "field yang tidak di-quote. Simpan ulang sebagai .xlsx atau perbaiki "
            f"quoting sebelum upload. Detail: {exc}"
        ) from exc


def parse_bom(file_bytes: bytes, filename: str) -> list[BomItemDraft]:
    """Parse a BOM file from raw bytes + the original filename.

    Filename suffix decides the format: `.xlsx` → openpyxl, otherwise CSV.
    """
    if not file_bytes:
        raise BomParseError("File BOM kosong.")

    suffix = Path(filename).suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        rows = _load_xlsx(file_bytes)
    else:
        rows = _load_csv(file_bytes)

    rows = [r for r in rows if any(c is not None and str(c).strip() != "" for c in r)]
    if not rows:
        raise BomParseError("File BOM tidak punya baris data.")

    string_rows: list[list[str]] = [
        [str(c) if c is not None else "" for c in row] for row in rows
    ]

    header_idx, mapping = _detect_header(string_rows)
    header_row = string_rows[header_idx]
    body = rows[header_idx + 1 :]

    drafts: list[BomItemDraft] = []
    for row in body:
        # pad short rows so col_idx lookups don't go off the end
        padded = list(row) + [None] * (len(header_row) - len(row))
        drafts.extend(_row_to_drafts(padded, mapping, header_row))
    return drafts
